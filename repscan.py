import socket
import openpyxl
import threading
import queue
import requests
from optparse import OptionParser  # 自定义输入参数
import time
import os
import re
import urllib3

list = []


class DoRun(threading.Thread):  # 自定义 多线程运行时使用的类
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self._queue = queue

    def run(self):
        while not self._queue.empty():
            date = req(self._queue.get())
            # print(date)
            if (date):
                list.append(date)


def Save_Data(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def check_http_https(ip, port):
    protocols = ['http', 'https']
    for protocol in protocols:
        try:
            url = f"{protocol}://{ip}:{port}"
            try:
                response = requests.get(url, timeout=3)
            except:
                continue
            if response.status_code == 200:
                return protocol
        except requests.ConnectionError:
            continue
    return "UNKNOWN"


def get_title(res, thread_count):  # 使用多线程 调用req ,获取datas(全局变量)
    que = queue.Queue()
    threads = []
    for date in res:
        url = ''
        for i in range(len(date['port'])):
            check = check_http_https(date['ip'], date['port'][i])
            if check == 'https':
                url = "https://" + date['ip'] + ":" + date['port'][i]
            elif check == 'http':
                url = "http://" + date['ip'] + ":" + date['port'][i]

        if url != '':
            que.put(url)
    for i in range(thread_count):
        threads.append(DoRun(que))  # 使用多线程 默认调用 run()函数
    for i in threads:
        i.start()  # 启动多线程
    for i in threads:
        i.join()  # 等待线程结束  后将数据保存至文件

    Save_title(list, str(int(time.time())))


def Save_title(data, filename):  # 将获取的title 保存到execle 文件中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'url'
    sheet['B1'] = 'title'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def req(url):  # 对域名进行验证，返回状态码，title
    heads = {  # 全局变量  请求头
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
        # 模拟浏览器请求
        'Connection': 'close',
        'Accept-Encoding': 'gzip, deflate'
    }
    data = {'url': '', 'title': ''}
    try:
        urllib3.disable_warnings()
        resp = requests.get(url=url, headers=heads, verify=False, timeout=10)  # 请求漏洞的url
        if resp.status_code == 200:
            title = re.findall("<title>(.*?)</title>", resp.text)
            data['url'] = url
            data['title'] = title
            print("[+]请求成功{}".format(data))
            return data
        else:
            print('[-]请求失败')
    except Exception as e:
        print('[-]请求失败e:')


# 用ping去扫描端口
def check_port(ip, ports):
    data = {'ip': ip, "port": []}
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    count = 0
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"ping {ip} -n {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
            count += 1
    if count == 0:
        print('Port {} of server {} has not been opened'.format(ports, ip))
        return {'ip': '', 'port': []}
    return data


# 常规扫描端口
def port_scan(ip, ports):  # 对域名进行验证，返回状态码，title
    host = ip
    data = {'ip': host, "port": []}
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    count = 0
    for port in range(int(port_start), int(port_end) + 1):
        sk = socket.socket()
        sk.settimeout(0.1)
        conn_result = sk.connect_ex((host, port))
        if conn_result == 0:
            print('Port {} of server {} has been opened'.format(port, host))
            data['port'].append(port)
            count += 1
        sk.close()
    if count == 0:
        print('Port {} of server {} has not been opened'.format(ports, ip))
        return {'ip': '', 'port': []}
    return data


# 获取用户输入的参数
def get_Input():
    optParser = OptionParser()
    optParser.add_option("-T", "--Type", action="store", type="int",
                         help='当Type的值为1 时扫描端口， 当Type的值为2时，扫描http服务', default=1)
    optParser.add_option("-t", "--thread_count", action="store", type="int", help='线程数量，默认为50', default=50)
    optParser.add_option("-i", "--ip", action="store", type="string", help='输入IP地址', default='None')
    optParser.add_option("-p", "--port", action="store", type="string", help='输入端口', default='0-65536')
    optParser.add_option("-f", "--filename", action="store", type="string", help='输入文件名(文件类型为xlsx)',
                         default='None')
    optParser.add_option("-r", "--iplist", action="store", type="string", help='输入ip文档', default='None')
    optParser.add_option("--sn", action="store_true", help='用ping扫描端口')
    (options, args) = optParser.parse_args()
    return options.Type, options.thread_count, options.ip, options.port, options.filename, options.iplist, options.sn


def run():
    Type, thread_count, ip, port, filename, file_ip, sn = get_Input()  # 获取用户输入的参数
    res = []
    if ip == 'None' and file_ip == 'None':
        print("无效ip")
        return

    if ip != 'None' and file_ip == 'None':
        if sn:
            res.append(check_port(ip, port))  # 调用check_port函数
        else:
            res.append(port_scan(ip, port))  # 进行端口扫描
    if file_ip != 'None':
        ip_list = []
        with open(file_ip, 'r') as f:
            line = f.readline()
            while line:
                if "\n" in line:
                    ip_list.append(line.split("\n")[0])
                else:
                    ip_list.append(line)
                line = f.readline()
        for i in ip_list:
            if sn:
                res.append(check_port(i, port))  # 调用check_port函数
            else:
                res.append(port_scan(i, port))  # 进行端口扫描
    if filename != 'None' and Type == 1:  # 将数据保存到xlsx
        Save_Data(res, filename)
    if Type == 2:
        get_title(res, thread_count)

run()
