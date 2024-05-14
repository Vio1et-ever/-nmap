import nmap
from openpyxl import *  # 数据处理，将获取到的数据保存在excel文件中
import threading
import queue
import urllib3
import requests
from optparse import OptionParser  # 自定义输入参数
import time
import re

numb_req = 0
list = []


class DoRun(threading.Thread):  # 自定义 多线程运行时使用的类
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self._queue = queue

    def run(self):
        while not self._queue.empty():
            date = req(self._queue.get())
            # print(date)
            if (date):
                list.append(date)


def init_excel(filename, sheetName):  # 创建.xlsx表格，并初始化内容
    wb = Workbook()
    if (sheetName == "PortScan"):
        head = ['numb', 'ip', 'port', 'protocol', 'state']
    else:
        head = ['numb', 'url', 'title']
    ws = wb.create_sheet(sheetName, index=0)
    for i in range(0, len(head)):
        ws.cell(1, i + 1).value = head[i]
    wb.save(filename)


def Save_Data(datas, filename):  # 将数据存储到表格当中
    filename = filename + ".xlsx"
    init_excel(filename, "PortScan")
    wb_save = load_workbook(filename)
    ws_save = wb_save.worksheets[0]
    for data in datas:
        print(data)
        current_col = 1
        for key in data:
            ws_save.cell(data['numb'] + 1, current_col).value = str(data[key])
            current_col += 1
    wb_save.save(filename)


def get_datas(res):  # 将 nmap 返回的数据 进行处理，返回 list[dir]
    dir = res['scan']
    numb = 0
    datas = []
    for k in dir.keys():
        # ip=(dir[k]['addresses']['ipv4'])
        for p in dir[k]['tcp']:
            data = {'numb': '', 'ip': '', 'port': '', 'protocol': '', 'state': ''}
            numb += 1
            data['numb'] = numb
            data['ip'] = str(dir[k]['addresses']['ipv4'])
            data['port'] = str(p)
            data['state'] = dir[k]['tcp'][p]['state']
            data['protocol'] = dir[k]['tcp'][p]['name']
            # print("[+]data={}".format(data))
            datas.append(data)
    return datas


def get_title(datas, thread_count):  # 使用多线程 调用req ,获取datas(全局变量)
    que = queue.Queue()
    threads = []
    for date in datas:
        url = ''
        if (date['protocol'] == 'https'):
            url = "https://" + date['ip'] + ":" + date['port']
        elif ("http" in date['protocol']):
            url = "http://" + date['ip'] + ":" + date['port']

        if (url != ''):
            que.put(url)
    for i in range(thread_count):
        threads.append(DoRun(que))  # 使用多线程 默认调用 run()函数
    for i in threads:
        i.start()  # 启动多线程
    for i in threads:
        i.join()  # 等待线程结束  后将数据保存至文件

    Save_title(list, str(int(time.time())))


def Save_title(datas, filename):  # 将获取的title 保存到execle 文件中
    filename = filename + ".xlsx"
    init_excel(filename, "title")
    wb_save = load_workbook(filename)
    ws_save = wb_save.worksheets[0]
    for data in datas:
        current_col = 1
        for key in data:
            ws_save.cell(data['numb'] + 1, current_col).value = str(data[key])
            current_col += 1
    wb_save.save(filename)


def req(url):  # 对域名进行验证，返回状态码，title
    global numb_req
    heads = {  # 全局变量  请求头
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36',
        # 模拟浏览器请求
        'Connection': 'close',
        'Accept-Encoding': 'gzip, deflate'
    }
    data = {'numb': '', 'url': '', 'title': ''}
    try:
        urllib3.disable_warnings()
        resp = requests.get(url=url, headers=heads, verify=False, timeout=10)  # 请求漏洞的url
        if resp.status_code == 200:
            title = re.findall("<title>(.*?)</title>", resp.text)
            numb_req += 1
            data['numb'] = numb_req
            data['url'] = url
            data['title'] = title
            # print("[+]"+url+"\ttitle:"+title)
            print("[+]请求成功{}".format(data))
            return data
        else:
            print('[-]请求失败')
    except Exception as e:
        print('[-]请求失败e:')


def get_Input():  # 获取用户输入的参数  ，返回  argument,
    optParser = OptionParser()
    optParser.add_option('-a', '--arguments', action='store', type="string", help='调用nmap模块 使用的参数',
                         default="-iL r.txt")
    optParser.add_option("-T", "--Type", action="store", type="int",
                         help='当Type的值为1 时扫描端口， 当Type的值为2时，扫描http服务', default=1)
    optParser.add_option("-t", "--thread_count", action="store", type="int", help='线程数量，默认为50', default=50)
    (options, args) = optParser.parse_args()
    return options.arguments, options.Type, options.thread_count


def print_info(datas):
    for data in datas:
        print("[+]" + str(data['ip']) + "  " + str(
            data['port'] + "  " + str(data['protocol']) + "  is  " + str(data['state'])))


def run():
    arguments, Type, thread_count = get_Input()  # 获取用户输入的参数
    print("arguments={},Type={},thread_count={}".format(arguments, Type, thread_count))
    np = nmap.PortScanner()
    res = np.scan(hosts='', arguments=arguments)  # 调用nmap模块 进行端口扫描
    datas = get_datas(res)  # 将扫描到的数据进行整理，提取
    filename = str(int(time.time()))  # 文件名为时间戳

    if (Type == 1):
        Save_Data(datas, filename)  # 当Type=1时，直接扫描端口  然后保存
    elif (Type == 2):
        get_title(datas, thread_count)  # 当Type=2时，先扫描http服务，然后使用requests模块进行验证，获取title值，再保存


run()
