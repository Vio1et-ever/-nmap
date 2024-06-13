import os
import openpyxl
import asyncio
import aiohttp
import socket


async def check_port_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def check_port(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scan_ports(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_port_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def check_port_ping(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    count = 0
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"ping {ip} -n {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
            count += 1
    if count == 0:
        print('Port {} of server {} has not been opened'.format(ports, ip))
        data = {'ip': '', 'port': [], 'protocol': [], 'service': []}



def udp_port_scan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcp_port_scan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def get_service_name(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


def Save_Data(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))
