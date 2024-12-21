import os
import openpyxl
import asyncio
import aiohttp
import socket
import gevent
from gevent import monkey
import platform
import subprocess
from queue import Queue
import IPy
import re
from scapy.all import *


async def check_port_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def check_port(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scan_ports(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def check_port_ping(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udp_port_scan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcp_port_scan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def get_service_name(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def syn_scan(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)

ip_que = Queue()
string = []


def star_ping(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def ping_func():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def list_of_groups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def list_ip(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def get_survive_ip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def Save_Data_xlsx(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Save_data_txt(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')


async def check_port_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def check(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scan(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def checkping(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udp(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcp(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def service_name(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def syn(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)


def ping(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def func():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def groups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def ip_l(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def getip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def Savexlsx(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Savetxt(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')



async def port_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def general_port(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scans(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def checkp(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udpn(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcpn(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def gete(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def synn(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)



def stag(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def p_func():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def lgroups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def l_ip(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def geip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def Save_xlsx(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Savetxt(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')


async def check_p(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def checkt(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scanp(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def checkp(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udpscan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcpscan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def getname(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def syns(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)



def star_p(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def ping_f():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def listgroups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def listip(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def getsip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def SaveData_xlsx(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Savedata_txt(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')



async def checkport_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def checkport(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scanports(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def checkport_ping(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udpport_scan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcpport_scan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def getservice_name(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def synscan(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)



def starping(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def pingfunc():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def listof_groups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def lisip(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def getsurvive_ip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def Save_Dataxlsx(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Save_datatxt(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')



async def check_porthttp(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def checkpor(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scanport(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def check_portping(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udp_portscan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcp_portscan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def get_servicename(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def synsca(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)



def starpin(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def pingfun():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def list_ofgroups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def listip_list(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def get_surviveip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def SaveDataxlsx(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Savedatatxt(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')


async def ccheck_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def ccheck_port(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def sscanports(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)


def check_portp(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"telnet {ip} {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
        else:
            print('Port {} of server {} has not been opened'.format(ports, ip))
            data = {'ip': '', 'port': [], 'protocol': [], 'service': []}


def udp_scan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcp_scan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def get_s_name(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


# syn扫描
def syns(ip, ports, data):
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    for port in range(int(port_start), int(port_end) + 1):
        p = IP(dst=ip) / TCP(dport=int(port))
        ans = sr1(p, timeout=1, verbose=1)
        print(ans)
        if ans is not None and ans[TCP].flags == 'SA':
            print(ip, "port", port, "is open.")
            data['port'].append(port)
        else:
            print(ip, "port", port, "is closed.")


monkey.patch_all(thread=False)



def sping(ip_list):
    for ip in ip_list:
        ip_que.put(ip)
    # 开启多协程
    cos = []
    for i in range(len(ip_list)):
        # 调用工作函数
        c = gevent.spawn(ping_func)
        cos.append(c)
    gevent.joinall(cos)


def pinfun():
    while True:
        if ip_que.qsize() == 0:
            break
        ip = ip_que.get()
        if (platform.system() == 'Windows'):
            print('ping -n 1 {}'.format(ip))
            ping = subprocess.Popen(
                'ping -n 1 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        else:
            ping = subprocess.Popen(
                'ping -c 2 {}'.format(ip),
                shell=False,
                close_fds=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        try:
            out, err = ping.communicate(timeout=8)
            if 'ttl' in out.decode('GBK').lower():
                string.append("{}".format(ip))
        except:
            pass
        ping.kill()


def list_groups(init_list, childern_list_len):
    list_of_groups = zip(*(iter(init_list),) * childern_list_len)  # 使用zip函数将列表按照网段长度分成多个列表
    end_list = [list(i) for i in list_of_groups]  # 转换成列表
    count = len(init_list) % childern_list_len
    end_list.append(init_list[-count:]) if count != 0 else end_list
    return end_list


def liip(ip):
    newIplist = []
    if "/" in ip:
        if int(ip.split("/")[1]) >= 24:
            num = ip.split('/')[1]
            length = len(IPy.IP('127.0.0.0/{}'.format(num)))  # 计算网段的IP个数
            endiplists = list_of_groups(range(0, 256), length)  # 将整个C段按子网掩码划分成多个列表
            for endiplist in endiplists:  # 判断输入IP所在的子网
                if int(ip.split('/')[0].split('.')[-1].strip()) in endiplist:
                    for endip in endiplist:
                        newIplist.append(
                            '.'.join(ip.split('/')[0].split('.')[:-1]) + '.{}'.format(endip))  # 以.为连接符，组合IP。
                    break
        elif int(ip.split("/")[1]) >= 16:
            new_ip = ip.split(".")[0] + "." + ip.split(".")[1] + ".0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
        else:
            new_ip = ip.split(".")[0] + ".0.0.0/{}".format(ip.split("/")[1])
            ips = IPy.IP(new_ip)
            for i in ips:
                newIplist.append(str(i))
    elif re.match(r'^\d{0,3}.\d{0,3}.\d{0,3}.\d{0,3}$', ip) != None:
        newIplist.append(ip)
    return newIplist


def surviveip(ip):
    newIplist = list_ip(ip)
    star_ping(newIplist)


def SaveDatax(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def Save_datat(data, filename):
    file = filename + '.txt'
    with open(file, 'w') as fp:
        for s in data:
            fp.write(s + '\n')
