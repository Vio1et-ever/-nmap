import os
from optparse import OptionParser
import openpyxl
import asyncio
import aiohttp
import socket

res = []


async def check_port_http(ip, port, data):  # 扫描http
    url = f"http://{ip}:{port}"
    data['ip'] = url
    async with aiohttp.ClientSession() as session:
        try:
            async with session.get(url, timeout=3) as response:
                if response.status == 200:
                    print(f"Port {port} on {ip} is open")
                    data['protocol'].append('http')
                else:
                    print(f"Port {port} on {ip} is closed")
                    data['protocol'].append('')
        except Exception as e:
            print(f"Port {port} on {ip} failed to connect: {e}")


async def check_port(ip, port, data):  # 常规扫描
    conn = asyncio.open_connection(ip, port)
    data['ip'] = ip
    try:
        reader, writer = await asyncio.wait_for(conn, timeout=0.5)
        print(f"Port {port} of {ip} is open")
        data['port'].append(port)
        writer.close()
        await writer.wait_closed()
    except:
        pass


async def scan_ports(ip, port, type, data):  # 用异步时序端口扫描
    tasks = []
    ports = port
    try:
        port_start, port_end = map(int, ports.split('-'))
    except:
        port_start = port_end = int(ports)

    for port in range(port_start, port_end + 1):
        if type == 1:
            task = check_port(ip, port, data)
            tasks.append(task)
        elif type == 2:
            tasks.append(check_port_http(ip, port, data))

        if len(tasks) >= 500:
            await asyncio.gather(*tasks)
            tasks = []
    if tasks:
        await asyncio.gather(*tasks)
    # res.append(data)


def check_port_ping(ip, ports, data):  # 用ping扫描端口
    data['ip'] = ip
    try:
        port_start, port_end = ports.split('-')
    except:
        port_start = port_end = int(ports)
    count = 0
    for port in range(int(port_start), int(port_end) + 1):
        result = os.system(f"ping {ip} -n {port}")
        if result == 0:
            print('Port {} of server {} has been opened'.format(port, ip))
            data['port'].append(port)
            count += 1
    if count == 0:
        print('Port {} of server {} has not been opened'.format(ports, ip))
        data = {'ip': '', 'port': [], 'protocol': [], 'service': []}



def udp_port_scan(ip, port):  # udp扫描端口
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    sock.settimeout(1)  # 设置超时时间
    try:
        # 发送一个空的UDP数据包
        sock.sendto(b'', (ip, port))
        # 尝试接收响应
        data, _ = sock.recvfrom(1024)
        return 'udp'
    except socket.timeout:
        return False
    except socket.error:
        return False


def tcp_port_scan(host, port):  # 扫描TCP端口
    try:
        # 创建一个TCP socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.connect((host, port))  # 尝试连接
        return 'tcp'
    except socket.error as e:
        return


def get_service_name(port, protocol):  # 获取端口的服务
    try:
        # 获取端口对应的服务名称
        service_name = socket.getservbyport(port, protocol)
        return service_name
    except OSError:
        return 'unknown'
    except TypeError:
        return


def get_input():  # 获取输入参数
    parser = OptionParser()
    parser.add_option("-T", "--Type", action='store', type='int', help="T=1 scan port, T=2 scan http service",
                      default=1)
    parser.add_option("-i", "--ip", dest="ip", help="IP address to scan", default=None)
    parser.add_option("-p", "--ports", dest="ports", help="Port range to scan, e.g., '1-1024'", default="1-65536")
    parser.add_option("-f", '--filename', action='store', type='string', help="filename", default=None)
    parser.add_option("-r", "--ipfile", action='store', type='string', help="A file that contains an IP address",
                      default=None)
    parser.add_option("--sn", action='store_true', help='Scan the port with ping', default=None)
    parser.add_option("--su", action='store_true', help='Use UDP for scanning', default=None)
    options, _ = parser.parse_args()
    return options.ip, options.ports, options.filename, options.Type, options.ipfile, options.sn, options.su


def Save_Data(data, filename):  # 将数据存储到表格当中
    # 创建新工作簿
    workbook = openpyxl.Workbook()
    # 获取默认工作表
    sheet = workbook.active
    # 写入数据到单元格
    current_row = 2
    sheet['A1'] = 'ip'
    sheet['B1'] = 'port'
    sheet['C1'] = 'state'
    sheet['D1'] = 'protocol'
    sheet['E1'] = 'service'
    for i in range(len(data)):
        for j in range(len(data[i]['port'])):
            sheet.cell(row=current_row, column=1, value=data[i]['ip'])
            sheet.cell(row=current_row, column=2, value=data[i]['port'][j])
            sheet.cell(row=current_row, column=3, value='open')
            sheet.cell(row=current_row, column=4, value=data[i]['protocol'][j])
            sheet.cell(row=current_row, column=5, value=data[i]['service'][j])
            current_row += 1

    # 保存工作簿
    workbook.save('{}.xlsx'.format(filename))


def main():
    ip, port, filename, type, file_ip, sn, su = get_input()
    data = {'ip': '', 'port': [], 'protocol': [], 'service': []}
    if ip or file_ip:
        # 如果ip地址存在
        if ip:
            asyncio.run(scan_ports(ip, port, 1, data))
        # 如果文件有IP地址
        if file_ip:
            ip_list = []
            # 获取IP地址
            with open(file_ip, 'r') as f:
                line = f.readline()
                while line:
                    if "\n" in line:
                        ip_list.append(line.split("\n")[0])
                    else:
                        ip_list.append(line)
                    line = f.readline()
            for i in ip_list:
                if sn:
                    check_port_ping(i, port, data)  # 进行ping扫描
                    print(f"D: {data}")
                    res.append(data)
                    data = {'ip': '', 'port': [], 'protocol': [], 'service': []}
                    print(f"R: {res}")
                else:
                    asyncio.run(scan_ports(i, port, 1, data))  # type = 1 进行端口扫描 type = 2 进行http连接
                    res.append(data)
                    data = {'ip': '', 'port': [], 'protocol': [], 'service': []}
    else:
        print("The IP address entered is invalid")
        return
    for i in range(len(res)):
        for j in range(len(res[i]['port'])):
            res[i]['protocol'].append(tcp_port_scan(res[i]['ip'], res[i]['port'][j]))
            res[i]['service'].append(get_service_name(res[i]['port'][j], res[i]['protocol'][j]))
    if type == 2:
        for i in range(len(res)):
            for port in res[i]['port']:
                asyncio.run(scan_ports(res[i]['ip'], port, 2, data))
    if su:
        for i in range(len(res)):
            for j in range(len(res[i]['port'])):
                res[i]['protocol'][j] = udp_port_scan(res[i]['ip'], res[i]['port'][j])

    # 保存文件
    if filename:
        Save_Data(res, filename)


main()
